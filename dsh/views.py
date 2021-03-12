from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views import generic
from django.views.generic.base import TemplateView
import json
from dateutil import relativedelta

from datetime import datetime, timezone, timedelta
from django.db.models import Count, Max, Min, Sum, F, Q, When, Subquery
from django.db.models.functions import TruncMonth
from bases.views import SinPermisos
from trn.models import SolicitudTrabajo, OrdenTrabajo, DetalleOrden, DetalleSolicitud
from ctg.models import ConteoCotizaciones, Items, Categorias, \
    Colores
from cfg.models import ConfiguracionReporteUsuarios
from bases.models import Estados
from est.models import Tiendas, Zonas
from usr.models import UsuariosTiendas, Usuarios
from ubc.models import Ciudades
from .serializers import OrdenesSerializer
from ctg.serializers import CategoriasSerializer, ItemsSerializer, \
    ColoresSerializer
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

# reportes
from openpyxl import Workbook
from trn.resources import UsuariosResource



class DashboardAdminOpeAPIView(APIView):
    def get(self, request, opcion, f_inicio=None, f_fin=None):
        contexto = {}
        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        dia_uno_mes_anterior = dia_uno_mes_actual - timedelta(days=1)
        dia_uno_mes_anterior = dia_uno_mes_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        if opcion == 1:
            fecha_inicio = dia_uno_mes_actual 
            fecha_fin = ahora
        if opcion == 2:
            fecha_inicio = dia_uno_mes_anterior
            fecha_fin = dia_uno_mes_actual
        if opcion == 3:

            fecha_inicio = datetime.strptime(f_inicio, "%Y-%m-%d")
            fecha_fin = datetime.strptime(f_fin, "%Y-%m-%d")
        #  ordenes por estado
        ordenes = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial()
            ).\
            order_by('estado').\
            values('estado').\
            annotate(cont=Count('estado'))
        totales_por_estado = []
        for orden in ordenes:
            estado = Estados.objects.\
                filter(id_estado=orden.get('estado')).first()
            orden_tmp = [
                str(estado.descripcion),
                str(orden.get('cont'))
            ]
            totales_por_estado.append(orden_tmp)

        totales_por_estado = set(map(tuple, totales_por_estado))

        totales_por_estado = sorted(
            totales_por_estado,
            key=lambda orden: int(orden[1]),
            reverse=True
        )

        totales_por_estado = json.dumps(totales_por_estado)
        
        #  producto mas cotizado
        cotizaciones = ConteoCotizaciones.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                usuario__pais=self.request.user.pais,
                item__tipo_catalogo__descripcion='PRODUCTOS'

            ).\
            order_by('item__taller', 'item_id').\
            values('item__taller', 'item_id').\
            annotate(cont=Count('item__taller'))
        
        if cotizaciones:
            maximo = cotizaciones.order_by('-cont')[:1]
            producto_mas_cotizado = Items.objects.\
                filter(id_item=maximo[0].get('item_id')).first()
            producto_mas_cotizado = ItemsSerializer(producto_mas_cotizado).data
        else:
            producto_mas_cotizado = None

        #servicio mas solicitado
        items_servicio = Items.objects.filter(tipo_catalogo__descripcion='SERVICIOS', taller__pais=self.request.user.pais).values('id_item')

        detalle_solcitado = DetalleSolicitud.objects.\
            filter(
                item_id__in=items_servicio
            ).\
            values('solicitud_id')
        solicitudes_id = SolicitudTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                pais=self.request.user.pais,
                id_solicitud__in=detalle_solcitado

            ).\
            exclude(sku_relacionado=None).\
            order_by('-sku_relacionado').\
            values('sku_relacionado').\
            annotate(contador=Count('sku_relacionado'))

        if solicitudes_id:
            maximo = solicitudes_id.order_by('-contador')[:1]
            servicio_mas_solicitado = Items.objects.\
                filter(
                    id_item=maximo[0].get('sku_relacionado')
                ).first()
            servicio_mas_solicitado = ItemsSerializer(servicio_mas_solicitado).data
        else:
            servicio_mas_solicitado = None

        #  relacion de ordenes internas/externas
        ordenes = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
            ).\
            order_by('externa').\
            values('externa').\
            annotate(cont=Count('externa'))
        totales_int_ext = []
        for orden in ordenes:
            externa = orden.get('externa')
            if externa:
                descripcion = 'Externas'
            else:
                descripcion = 'Internas'
            orden_tmp = [
                descripcion,
                str(orden.get('cont'))
            ]
            totales_int_ext.append(orden_tmp)
        totales_int_ext = set(map(tuple, totales_int_ext))
        totales_int_ext = sorted(
            totales_int_ext,
            key=lambda orden: int(orden[1]),
            reverse=True
        )
        totales_int_ext = json.dumps(totales_int_ext)
        

        #  relacion de solicitudes internas/externas
        solicitudes = SolicitudTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
            ).\
            order_by('externa').\
            values('externa').\
            annotate(cont=Count('externa'))
        tot_int_ext = []
        for solicitud in solicitudes:
            externa = solicitud.get('externa')
            if externa:
                descripcion = 'Externas'
            else:
                descripcion = 'Internas'
            slc_tmp = [
                descripcion,
                str(solicitud.get('cont'))
            ]
            tot_int_ext.append(slc_tmp)
        tot_int_ext = set(map(tuple, tot_int_ext))
        tot_int_ext = sorted(
            tot_int_ext,
            key=lambda solicitud: int(solicitud[1]),
            reverse=True
        )
        tot_int_ext = json.dumps(tot_int_ext)

        # total de gramos fabricados
        
        gramos_fabricados = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                pais=self.request.user.pais,
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS'
            ).\
            values('peso_final').\
            aggregate(suma=Sum('peso_final'))


        #  producto mas fabricado
        items_producto = Items.objects.filter(tipo_catalogo__descripcion='PRODUCTOS', taller__pais=self.request.user.pais).values('id_item')
        
        ord_id = OrdenTrabajo.objects.filter(pais=self.request.user.pais).values('id_orden')
        items_id = DetalleOrden.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                pais=self.request.user.pais,
                orden_id__in=ord_id,
                id_item__in=items_producto
            ).\
            order_by('id_item').\
            values('id_item').\
            annotate(cont=Count('id_item'))

        if items_id:
            mayor = items_id.order_by('-cont')[:1]
            producto_mas_fabricado = Items.objects.filter(
                id_item=mayor[0].get('id_item')
            ).first()
            producto_mas_fabricado = ItemsSerializer(producto_mas_fabricado).data
        else:
            producto_mas_fabricado = None



        #  ordenes por mes
        ordenes_mes = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=self.request.user.grupo_empresarial(),
                categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
            order_by('mes').\
            values('mes').\
            annotate(suma=Sum('peso_final'))
        

        fabricaciones_por_mes = []

        for orden in ordenes_mes:
            orden_tmp = [
                str(orden.get('mes')),
                str(orden.get('suma'))
            ]

            if orden.get('mes') != None:
                fabricaciones_por_mes.append(orden_tmp)
        fabricaciones_por_mes = json.dumps(fabricaciones_por_mes)


        # categoria mas fabricada
        ordenes = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                pais=request.user.pais,
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS'
            ).\
            order_by('categoria').\
            values('categoria').\
            annotate(cont=Count('categoria'))
        if ordenes:
            mayor = ordenes.order_by('-cont')[:1]
            categoria_mas_fabricada = Categorias.objects.\
                filter(id_categoria=mayor[0].get('categoria')).first()
            categoria_mas_fabricada = \
                CategoriasSerializer(categoria_mas_fabricada).data
        else:
            categoria_mas_fabricada = None

        # color mas fabricado
        items_color = Items.objects.filter(taller__pais=self.request.user.pais).values('id_item')
        ord_color = OrdenTrabajo.objects.filter(
                    fecha_creacion__gte=fecha_inicio,
                    fecha_creacion__lte=fecha_fin,
                    pais=self.request.user.pais,
                    categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
                    ).\
                    exclude(color__descripcion='SIN COLOR').\
                    values('color').\
                    order_by('color').\
                    annotate(contador=Count('color'))
        if ord_color:
            mayor = ord_color.order_by('-contador')[:1]
            color_mas_fabricado = Colores.objects.\
                filter(id_color=mayor[0].get('color')).first()
            color_mas_fabricado = ColoresSerializer(color_mas_fabricado).data
            
        else:
            color_mas_fabricado = None

        # ranking joyerias de gramos fabricados
        ordenes_mes = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=self.request.user.grupo_empresarial(),
                categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            order_by('tienda__nombre').\
            values('tienda__nombre').\
            annotate(suma=Sum('peso_final')).\
            order_by('-suma')[:10]
        
        gramos_por_joyeria = []
        for orden in ordenes_mes:
            orden_tmp = [
                str(orden.get('tienda__nombre')),
                str(orden.get('suma'))
            ]
            gramos_por_joyeria.append(orden_tmp)
        gramos_por_joyeria = json.dumps(gramos_por_joyeria)
        

        #ranking de las tiendas que generan mas ordenes
        ordenes_tiendas = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial()
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            values('tienda__nombre').\
            annotate(contador=Count('tienda_id')).order_by('-contador')[:5]

        ordenes_por_tienda = []            
        for i in ordenes_tiendas:
            tienda_contador = [
                str(i.get('tienda__nombre')),
                str(i.get('contador'))
            ]
            ordenes_por_tienda.append(tienda_contador)
        ordenes_por_tienda = json.dumps(ordenes_por_tienda)

        #total de solicitudes rechazadas
        tot_sol_rech = SolicitudTrabajo.objects.\
            filter(
                pais=self.request.user.pais,
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                estado_id=5
            ).\
            values('id_solicitud').\
            aggregate(contador=Count('id_solicitud'))


        # trabajado en reunion con Julio Moran
        ordenes_obj = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial()
            ).\
            values('id_orden')

        detalle = DetalleOrden.objects.\
            filter(
                orden_id__in=ordenes_obj,
                id_solicitud=None
            ).\
            values('id_item')

        items_catalogo = Items.objects.\
            filter(
                id_item__in=detalle,
                tipo_catalogo_id=1
            ).values('id_item')
        
        items_servicio = Items.objects.\
            filter(
                id_item__in=detalle,
                tipo_catalogo_id=2
            ).values('id_item')

        ordenes_catalogo = DetalleOrden.objects.\
            filter(
                orden_id__in=ordenes_obj,
                id_item__in=items_catalogo
            )
        ordenes_servicio = DetalleOrden.objects.\
            filter(
                orden_id__in=ordenes_obj,
                id_item__in=items_servicio
            )
        

        # ranking de ciudades con mas ordenes_mes
        ciudades_id = Ciudades.objects.filter(pais_id=self.request.user.pais_id).values('id_ciudad')

        ordenes_ciudad = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__ciudad_id__in=ciudades_id,
                pais=self.request.user.pais
            ).values('tienda__ciudad__nombre').\
            annotate(contador=Count('id_orden')).\
            order_by('-contador')[:1]

        orden_ciudad = {}
        for ciudad in ordenes_ciudad:
            orden_ciudad={
                'ciudad': ciudad.get('tienda__ciudad__nombre'),
                'contador': ciudad.get('contador')
            }

        orden_ciudad = json.dumps(orden_ciudad)

        ''' zona con mas ordenes '''
        zonas_id = Zonas.objects.filter(grupo_empresarial=self.request.user.grupo_empresarial()).values('id_zona')

        ordenes_zona = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__zona_id__in=zonas_id,
                pais=self.request.user.pais
            ).values('tienda__zona__nombre').\
            annotate(contador=Count('id_orden')).\
            order_by('-contador')[:1]
        orden_zona = {}
        for zona in ordenes_zona:
            orden_zona={
                'zona': zona.get('tienda__zona__nombre'),
                'contador': zona.get('contador')
            }
        #     lista_zona.append(orden_tmp)

        orden_zona = json.dumps(orden_zona)

        contexto = {
                    # 'ordenes': ordenes,
                    'totales_por_estado': totales_por_estado,
                    'producto_mas_cotizado': producto_mas_cotizado,
                    'servicio_mas_solicitado': servicio_mas_solicitado,
                    'totales_int_ext': totales_int_ext,
                    'tot_int_ext': tot_int_ext,
                    'gramos_fabricados': gramos_fabricados.get('suma'),
                    'producto_mas_fabricado': producto_mas_fabricado,
                    'fabricaciones_por_mes': fabricaciones_por_mes,
                    'categoria_mas_fabricada': categoria_mas_fabricada,
                    'color_mas_fabricado': color_mas_fabricado,
                    'gramos_por_joyeria': gramos_por_joyeria,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin,
                    'ordenes_por_tienda': ordenes_por_tienda,
                    'total_solicitud_rechazadas': tot_sol_rech.get('contador'),
                    'orden_ciudad': orden_ciudad,
                    'orden_zona': orden_zona
                    }
        return JsonResponse(contexto)



class DashboardAdminTalAPIView(APIView):
    def get(self, request, opcion, f_inicio=None, f_fin=None):
        contexto = {}

        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        dia_uno_mes_anterior = dia_uno_mes_actual - timedelta(days=1)
        dia_uno_mes_anterior = dia_uno_mes_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        if opcion == 1:
            fecha_inicio = dia_uno_mes_actual
            fecha_fin = ahora
        if opcion == 2:
            fecha_inicio = dia_uno_mes_anterior
            fecha_fin = dia_uno_mes_actual
        if opcion == 3:
            fecha_inicio = f_inicio
            # print(datetime.strptime(f_fin, '%Y-%m-%d') +timedelta(days=1))
            fecha_fin = datetime.strptime(f_fin, '%Y-%m-%d') +timedelta(days=1)
        #  ordenes por estado
        ordenes = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                taller=request.user.taller()
            ).\
            order_by('estado').\
            values('estado').\
            annotate(cont=Count('estado'))

        totales_por_estado = []
        for orden in ordenes:
            estado = Estados.objects.\
                filter(id_estado=orden.get('estado')).first()
            orden_tmp = [
                str(estado.descripcion),
                str(orden.get('cont'))
            ]
            totales_por_estado.append(orden_tmp)

        totales_por_estado = set(map(tuple, totales_por_estado))
        totales_por_estado = sorted(
            totales_por_estado,
            key=lambda orden: int(orden[1]),
            reverse=True
        )
        totales_por_estado = json.dumps(totales_por_estado)

        #  producto mas cotizado
        item_productos_id = Items.objects.filter(tipo_catalogo__descripcion='PRODUCTOS').values('id_item')
        cotizaciones = ConteoCotizaciones.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                item_id__in=item_productos_id,
                item__taller=request.user.taller()
            ).\
            order_by('item__taller', 'item_id').\
            values('item__taller', 'item_id').\
            annotate(cont=Count('item__taller'))
        if cotizaciones:
            maximo = cotizaciones.order_by('-cont')[:1]
            producto_mas_cotizado = Items.objects.\
                filter(id_item=maximo[0].get('item_id')).first()
            producto_mas_cotizado = ItemsSerializer(producto_mas_cotizado).data
            
        else:
            producto_mas_cotizado = None

        
        # Solicitudes por estado
        solicitudes = SolicitudTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                taller=request.user.taller()
            ).\
            order_by('estado').\
            values('estado').\
            annotate(cont=Count('estado'))
        solicitudes_por_estado = []
        for solicitud in solicitudes:
            estado = Estados.objects.\
                filter(id_estado=solicitud.get('estado')).first()
            solicitud_tmp = [
                str(estado.descripcion),
                str(solicitud.get('cont'))
            ]
            solicitudes_por_estado.append(solicitud_tmp)
        solicitudes_por_estado = set(map(tuple, solicitudes_por_estado))
        solicitudes_por_estado = sorted(
            solicitudes_por_estado,
            key=lambda orden: int(orden[1]),
            reverse=True
        )
       
        solicitudes_por_estado = json.dumps(solicitudes_por_estado)
        
        # total de gramos fabricados en productos
        gramos_fabricados = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                pais=self.request.user.pais
            ).\
            values('peso_final').\
            aggregate(suma=Sum('peso_final'))
        # print('gramos_fabricados')
        # print(gramos_fabricados)


        #total de gramos utilizado en servicios
        gramos_servicios = OrdenTrabajo.objects.\
            filter(
                categoria__division__tipo_catalogo__descripcion='SERVICIOS',
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                taller=self.request.user.taller(),
                pais=self.request.user.pais
                ).\
                values('peso_solicitado').\
                aggregate(sumat=Sum('peso_solicitado'))
        # print('gramos en slicitudes')
        # print(gramos_servicios)
        
        #  producto mas fabricado
        items = DetalleOrden.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                pais=request.user.pais,
                orden__categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                orden__taller=self.request.user.taller()
            ).\
            exclude(id_item=None).\
            order_by('id_item').\
            values('id_item').\
            annotate(cont=Count('id_item'))
        if items:
            mayor = items.order_by('-cont')[:1]
            producto_mas_fabricado = Items.objects.\
                filter(id_item=mayor[0].get('id_item')).first()
            producto_mas_fabricado = \
                ItemsSerializer(producto_mas_fabricado).data
        else:
            producto_mas_fabricado = None


        ''' obtener el total de gramos fabricados por mes en las ordenes Productos'''
        ordenes_mes = OrdenTrabajo.objects.\
            filter(
                taller=self.request.user.taller(),
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
            order_by('mes').\
            values('mes').\
            annotate(suma=Sum('peso_final'))
        fabricaciones_por_mes = []


        # for orden in ordenes_mes:
        #     suma = orden.get('suma')
        #     if orden.get('mes') != None:
        #         if orden.get('suma') == None:
        #             suma = 0
        #             orden_tmp = [
        #                 str(orden.get('mes')),
        #                 str(suma)
        #             ]
        #         else:
        #             orden_tmp = [
        #                 str(orden.get('mes')),
        #                 str(suma)
        #             ]
        #         fabricaciones_por_mes.append(orden_tmp)
        # fabricaciones_por_mes = json.dumps(fabricaciones_por_mes)




        ordenes_mes = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__sociedad__grupo_empresarial=self.request.user.grupo_empresarial(),
                categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
            order_by('mes').\
            values('mes').\
            annotate(suma=Sum('peso_final'))
        

        fabricaciones_por_mes = []

        for orden in ordenes_mes:
            orden_tmp = [
                str(orden.get('mes')),
                str(orden.get('suma'))
            ]

            if orden.get('mes') != None:
                fabricaciones_por_mes.append(orden_tmp)
        fabricaciones_por_mes = json.dumps(fabricaciones_por_mes)




        ''' obtener el total de gramos en ordenes de servicios por mes'''
        servicios_mes = OrdenTrabajo.objects.\
            filter(
                categoria__division__tipo_catalogo__descripcion='SERVICIOS',
                taller=self.request.user.taller()
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
            order_by('mes').\
            values('mes').\
            annotate(suma=Sum('peso_solicitado'))
        gramos_servicios_mes = []

        for orden in servicios_mes:
            orden_tmp = [
                str(orden.get('mes')),
                str(orden.get('suma'))
            ]
            gramos_servicios_mes.append(orden_tmp)
        gramos_servicios_mes = json.dumps(gramos_servicios_mes)
        
        # categoria mas fabricada
        ordenes = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                pais=request.user.pais,
                taller=request.user.taller()
            ).\
            order_by('categoria').\
            values('categoria').\
            annotate(cont=Count('categoria'))
        if ordenes:
            mayor = ordenes.order_by('-cont')[:1]
            categoria_mas_fabricada = Categorias.objects.\
                filter(id_categoria=mayor[0].get('categoria')).first()
            categoria_mas_fabricada = \
                CategoriasSerializer(categoria_mas_fabricada).data
        else:
            categoria_mas_fabricada = None
        # color mas fabricado
        ord_color = OrdenTrabajo.objects.\
            filter(
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                taller=self.request.user.taller(),
                pais=self.request.user.pais
            ).\
            values('color_id').\
            order_by('color_id').\
            annotate(contador=Count('color_id'))

        if ord_color:
            mayor = ord_color.order_by('-contador')[:1]
            color_mas_fabricado = Colores.objects.\
                filter(id_color=mayor[0].get('color_id')).first()
            color_mas_fabricado = ColoresSerializer(color_mas_fabricado).data
            
        else:
            color_mas_fabricado = None

        # ranking joyerias
        
        ordenes_mes = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                pais=self.request.user.pais
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            order_by('tienda__nombre').\
            values('tienda__nombre').\
            annotate(suma=Sum('peso_final')).\
            order_by('-suma')[:5]

        gramos_por_joyeria = []
        for orden in ordenes_mes:
            orden_tmp = [
                str(orden.get('tienda__nombre')),
                str(orden.get('suma'))
            ]
            gramos_por_joyeria.append(orden_tmp)
        gramos_por_joyeria = json.dumps(gramos_por_joyeria)

        #ranking de gramos solicitados en servicios por tiendas
        servicios_tiendas = OrdenTrabajo.objects.\
            filter(
                categoria__division__tipo_catalogo__descripcion='SERVICIOS',
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                taller=self.request.user.taller(),
                pais=self.request.user.pais
            ).\
            values('tienda__nombre').\
            annotate(suma=Sum('peso_final')).order_by('-suma')[:5]

        
        servicios_por_tienda = []            
        for i in servicios_tiendas:
            tienda_contador = [
                str(i.get('tienda__nombre')),
                str(i.get('suma'))
            ]
            servicios_por_tienda.append(tienda_contador)
        servicios_por_tienda = json.dumps(servicios_por_tienda)

        #total de solicitudes rechazadas
        tot_sol_rech = SolicitudTrabajo.objects.\
            filter(
                pais=self.request.user.pais,
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                estado_id=5
            ).\
            values('id_solicitud').\
            aggregate(contador=Count('id_solicitud'))

        #servicios con mas ordenes
        servicio_orden = DetalleOrden.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin,
                orden__categoria__division__tipo_catalogo__descripcion='SERVICIOS'
            ).\
            values('id_item').\
            order_by('id_item').\
            annotate(contador=Count('id_item')).order_by('-contador')[:5]
        servicio_mas_ordenes = []
        for i in servicio_orden:
            servicio_name = Items.objects.filter(id_item=i.get('id_item')).values('id_item', 'descripcion')
            for j in servicio_name:
                if j['id_item'] == i.get('id_item'):
                    servicio_mas_ordenes.append([j['descripcion'], str(i.get('contador'))])

        # rankin de tienda que generan menos ordenes
        ordenes_tiendas = OrdenTrabajo.objects.\
            filter(
                fecha_creacion__gte=fecha_inicio,
                fecha_creacion__lte=fecha_fin
            ).\
            exclude(fecha_fin_trabajo__lte='2000-01-01').\
            values('tienda__nombre').\
            annotate(contador=Count('tienda_id')).order_by('contador')[:5]

        ordenes_por_tienda = []            
        for i in ordenes_tiendas:
            tienda_contador = [
                str(i.get('tienda__nombre')),
                str(i.get('contador'))
            ]
            ordenes_por_tienda.append(tienda_contador)
        ordenes_por_tienda = json.dumps(ordenes_por_tienda)
        # print(ordenes_tiendas)

        contexto = {
                    # 'ordenes': ordenes,
                    'totales_por_estado': totales_por_estado,
                    'producto_mas_cotizado': producto_mas_cotizado,
                    'solicitudes_por_estado': solicitudes_por_estado,
                    'gramos_fabricados': gramos_fabricados.get('suma'),
                    'producto_mas_fabricado': producto_mas_fabricado,
                    'fabricaciones_por_mes': fabricaciones_por_mes,
                    'categoria_mas_fabricada': categoria_mas_fabricada,
                    'color_mas_fabricado': color_mas_fabricado,
                    'gramos_por_joyeria': gramos_por_joyeria,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin,
                    'servicios_por_tienda': servicios_por_tienda,
                    'gramos_servicios_mes': gramos_servicios_mes,
                    'gramos_servicios': gramos_servicios.get('sumat'),
                    'total_solicitud_rechazadas': tot_sol_rech.get('contador'),
                    'servicio_mas_ordenes': servicio_mas_ordenes,
                    'tiendas_menos_ordenes': ordenes_por_tienda
                    }
        return JsonResponse(contexto)


class DashboardAdminOpeView(SinPermisos, generic.TemplateView):
    permission_required = "trn.estadisticas_admin"
    template_name = "dsh/dashboard_operaciones.html"
    login_url = "bases:login"


class DashboardAdminTalView(SinPermisos, generic.TemplateView):
    permission_required = "trn.estadisticas_admin"
    template_name = "dsh/dashboard_taller.html"
    login_url = "bases:login"


class ReporteTiendasEstadisticasExcel(TemplateView):

    def get(self, request, f_inicio, f_fin, opcion):
        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        dia_uno_mes_anterior = dia_uno_mes_actual - timedelta(days=1)
        dia_uno_mes_anterior = dia_uno_mes_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        if opcion == 1:
            fecha_inicio = dia_uno_mes_actual
            fecha_fin = ahora
        if opcion == 2:
            fecha_inicio = dia_uno_mes_anterior
            fecha_fin = dia_uno_mes_actual -timedelta(days=1)
        administrador = self.request.user.rol.descripcion = 'ADMIN OPERACIONES'
        tiendas_id = Tiendas.objects.filter(sociedad__grupo_empresarial=self.request.user.grupo_empresarial())
        if administrador:
            if opcion==1:
                ordenes_tiendas = OrdenTrabajo.objects.filter(
                        fecha_creacion__gte=fecha_inicio,
                        fecha_creacion__lte=fecha_fin,
                        tienda__sociedad__grupo_empresarial=self.request.user.grupo_empresarial()
                    ).\
                    values('tienda__nombre').\
                    annotate(
                        contador=Count('tienda_id', filter=Q(categoria__division__tipo_catalogo__descripcion='PRODUCTOS')),
                        contadors=Count('tienda_id', filter=Q(categoria__division__tipo_catalogo__descripcion='SERVICIOS')),
                        sumap=Sum('peso_final', filter=Q(categoria__division__tipo_catalogo__descripcion='PRODUCTOS')),
                        sumas=Sum('peso_final', filter=Q(categoria__division__tipo_catalogo__descripcion='SERVICIOS')),
                    ).\
                    order_by('-tienda__nombre')
            if opcion==2:
                ordenes_tiendas = OrdenTrabajo.objects.filter(
                        fecha_creacion__gte=fecha_inicio,
                        fecha_creacion__lte=fecha_fin,
                        tienda__sociedad__grupo_empresarial=self.request.user.grupo_empresarial()
                    ).\
                    values('tienda__nombre').\
                    annotate(
                        contador=Count('tienda_id', filter=Q(categoria__division__tipo_catalogo__descripcion='PRODUCTOS')),
                        contadors=Count('tienda_id', filter=Q(categoria__division__tipo_catalogo__descripcion='SERVICIOS')),
                        sumap=Sum('peso_final', filter=Q(categoria__division__tipo_catalogo__descripcion='PRODUCTOS')),#suma de gramos de orden prod.
                        sumas=Sum('peso_final', filter=Q(categoria__division__tipo_catalogo__descripcion='SERVICIOS')),#suma gramos de orden servic.

                    ).\
                    order_by('-tienda__nombre')
        if ordenes_tiendas:
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'REPORTE DE TIENDAS'
            ws.merge_cells('B1:E1')
            ws['B3'] = 'NOMBRE TIENDA'
            ws.merge_cells('B3:C3')
            ws['D3'] = 'TOTAL ORDENES PROD.'
            ws.merge_cells('D3:E3')
            ws['F3'] = 'GRAMOS PRODUCTOS'
            ws.merge_cells('F3:G3')

            ws['H3'] = 'TOTAL ORDENES SERV.'
            ws.merge_cells('H3:I3')

            ws['J3'] = 'GRAMOS SERVICIOS'
            ws.merge_cells('J3:K3')

            cont=4
            total_gramop = []
            total_gramos = []
            for i in ordenes_tiendas:
                ws.cell(row=cont, column=2, ).value = str(i.get('tienda__nombre'))
                ws.cell(row=cont, column=4).value = str(i.get('contador'))
                ws.cell(row=cont, column=6).value = str(i.get('sumap'))
                ws.cell(row=cont, column=8).value = str(i.get('contadors'))
                ws.cell(row=cont, column=10).value = str(i.get('sumas'))
                total_gramop.append(i.get('sumap'))
                total_gramos.append(i.get('sumas'))
                cont = cont + 1

            sumatotal = 0
            for i in total_gramop:
                if i == None:
                    i=0
                sumatotal = sumatotal + i
            
            sumatotal2 = 0
            for i in total_gramos:
                if i == None:
                    i=0
                sumatotal2 = sumatotal2 + i

            ws.cell(row=cont, column=5).value = str('total:')
            ws.cell(row=cont, column=6).value = str(sumatotal)
            ws.cell(row=cont, column=9).value = str('total:')
            ws.cell(row=cont, column=10).value = str(sumatotal2)

            
            nombre_archivo = 'Reporte_ordenes_' + ahora.strftime('%d-%m-%Y') + '.xls'
            response = HttpResponse(content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'attachment; filename=' + nombre_archivo
            wb.save(response)

        else:
            response  = HttpResponse('No existen ordenes para estas fechas.<a href="/">Regresar</a>')
        return response



class ReporteUsuariosEstadisticasExcel(TemplateView):

    def get(self, request, f_inicio, f_fin, opcion):
        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        dia_uno_mes_anterior = dia_uno_mes_actual - timedelta(days=1)
        dia_uno_mes_anterior = dia_uno_mes_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        if opcion == 1:
            fecha_inicio = dia_uno_mes_actual
            fecha_fin = ahora
        if opcion == 2:
            fecha_inicio = dia_uno_mes_anterior
            fecha_fin = dia_uno_mes_actual -timedelta(days=1)
        # administrador = self.request.user.rol.descripcion = 'ADMIN OPERACIONES'
        if self.request.user.tipo_usuario.descripcion == 'USUARIO OPERACIONES':
            roles_id = ConfiguracionReporteUsuarios.objects.\
                filter(
                    grupo_empresarial_id=self.request.user.grupo_empresarial().id_grupo_empresarial
                ).\
                values('rol_id').\
                order_by('first_name')
        else:
            roles_id = ConfiguracionReporteUsuarios.objects.\
                filter(
                    taller_id=self.request.user.taller().id_taller
                ).\
                values('rol_id').\
                order_by('first_name')

        # if administrador:
        if opcion==1:
            usuarios_conectados = Usuarios.objects.filter(
                    rol_id__in = roles_id,
                    pais=request.user.pais
                ).\
                values('id', 'username', 'first_name', 'last_name', 'date_joined', 'last_login', 'is_active', 'usuariostiendas__tienda')
            
        if opcion==2:
            usuarios_conectados = Usuarios.objects.filter(
                    rol_id__in = roles_id
                ).\
                    annotate(
                    tienda=F('usuariostiendas__tienda__nombre'),
                ).\
                values('username', 'first_name', 'last_name', 'date_joined', 'last_login', 'is_active', 'tienda')
        if opcion==3:
            usuarios_conectados = Usuarios.objects.filter(
                    rol_id__in = roles_id
                ).\
                    annotate(
                    tienda=F('usuariostiendas__tienda__nombre'),
                ).\
                order_by('tienda').\
                values('username', 'first_name', 'last_name', 'date_joined', 'last_login', 'is_active', 'tienda')

        if usuarios_conectados:
            usuarios_resource = UsuariosResource()
            dataset = usuarios_resource.export(usuarios_conectados)
            nombre_archivo = 'Reporte_conexion_usuarios_' + ahora.strftime('%d-%m-%Y') + '.xls'
            response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=' + nombre_archivo
        else:
            response = HttpResponse('No existen ordenes')
        return response




def gramos_fabricados(request):
    if request.is_ajax and request.method == "GET":
        usuario=get_object_or_404(Usuarios,id=request.user.id)
        estado=0
        print('********************')
        opcion = request.GET.get("opcion", None)
        anio_actual = datetime.now()
        anio_actual = anio_actual.year
        anio_anterior = (anio_actual) - 1
        dia_uno_anio_actual = datetime(anio_actual, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio = datetime(anio_actual, 12, 31, 11, 59, 59, 00000)
        dia_uno_anio_anterior = datetime(anio_anterior, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio_anterior = datetime(anio_anterior, 12, 31, 11, 59, 59, 00000)


        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        dia_uno_meses_anterior = dia_uno_mes_actual - relativedelta.relativedelta(months=6)
        dia_uno_meses_anterior = dia_uno_meses_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        dia_uno_meses_tres = dia_uno_mes_actual - relativedelta.relativedelta(months=3)
        dia_uno_meses_tres = dia_uno_meses_tres.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        if opcion == '1':
            fecha_inicio = dia_uno_anio_actual
            fecha_fin = ultimo_dia_anio
        if opcion == '2':
            fecha_inicio = dia_uno_anio_anterior
            fecha_fin = ultimo_dia_anio_anterior
        if opcion == '3':
            fecha_inicio = dia_uno_meses_anterior
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)
        if opcion == '4':
            fecha_inicio = dia_uno_meses_tres
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)

        
        if usuario.tipo_usuario.descripcion == 'USUARIO OPERACIONES':
            ordenes_mes = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial(),
                    categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
                ).\
                exclude(fecha_fin_trabajo__lte='2000-01-01').\
                extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
                extra({'anio': 'Extract(year from fecha_fin_trabajo)'}).\
                order_by('anio', 'mes').\
                values('mes').\
                annotate(suma=Sum('peso_final'))
        else:
            ordenes_mes = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    taller=request.user.taller(),
                    categoria__division__tipo_catalogo__descripcion="PRODUCTOS"
                ).\
                exclude(fecha_fin_trabajo__lte='2000-01-01').\
                extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
                extra({'anio': 'Extract(year from fecha_fin_trabajo)'}).\
                order_by('anio', 'mes').\
                values('mes', 'anio').\
                annotate(suma=Sum('peso_final'))


        for orden in ordenes_mes:
            if orden.get('suma') == None:
                orden['suma'] = float(0)
        
        fabricaciones_por_mes = []

        for orden in ordenes_mes:
            
            orden_tmp = [
                str(orden.get('mes')),
                str(orden.get('suma'))
            ]
            if orden.get('mes') != None:
                fabricaciones_por_mes.append(orden_tmp)

        cont = len(fabricaciones_por_mes)
        if opcion == '1':
            while cont <= 12:
                cont += 1
                orden_tmp = [
                    str(cont),
                    str(0.00)
                ]
                fabricaciones_por_mes.append(orden_tmp)
                if cont >= 12:
                    break
        cont2 = 1
        if opcion == '2':
            enumerar_meses = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
            fabricaciones_por_mes = sorted(
                fabricaciones_por_mes,
                key=lambda orden: int(orden[0]),
                reverse=False
            )
        # if opcion == '3':
        #     # enumerar_meses = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        #     print(fabricaciones_por_mes)
        #     # print(enumerar_meses)
        #     fabricaciones_por_mes = sorted(
        #         fabricaciones_por_mes,
        #         key=lambda orden: int(orden[0]),
        #         reverse=True
        #     )
        #     print(fabricaciones_por_mes)
                
        
        fabricaciones_por_mes = json.dumps(fabricaciones_por_mes)

        return JsonResponse(fabricaciones_por_mes, safe=False)
    return JsonResponse({}, status = 400)




def gramos_fabricados_servicios(request):
    if request.is_ajax and request.method == "GET":
        usuario=get_object_or_404(Usuarios,id=request.user.id)
        estado=0
        opcion = request.GET.get("opcion", None)
        anio_actual = datetime.now()
        anio_actual = anio_actual.year
        anio_anterior = (anio_actual) - 1
        dia_uno_anio_actual = datetime(anio_actual, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio = datetime(anio_actual, 12, 31, 11, 59, 59, 00000)
        dia_uno_anio_anterior = datetime(anio_anterior, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio_anterior = datetime(anio_anterior, 12, 31, 11, 59, 59, 00000)


        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        dia_uno_meses_anterior = dia_uno_mes_actual - relativedelta.relativedelta(months=6)
        dia_uno_meses_anterior = dia_uno_meses_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        # print(dia_uno_meses_anterior)

        dia_uno_meses_tres = dia_uno_mes_actual - relativedelta.relativedelta(months=3)
        dia_uno_meses_tres = dia_uno_meses_tres.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )


        if opcion == '1':
            fecha_inicio = dia_uno_anio_actual
            fecha_fin = ultimo_dia_anio
        if opcion == '2':
            fecha_inicio = dia_uno_anio_anterior
            fecha_fin = ultimo_dia_anio_anterior
        if opcion == '3':
            fecha_inicio = dia_uno_meses_anterior
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)
            fecha_fin = fecha_fin.replace(
                hour=11,
                minute=59,
                second=0
                )
        if opcion == '4':
            fecha_inicio = dia_uno_meses_tres
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)

        # print(fecha_inicio)
        # print(fecha_fin)

        # print('********')
        if usuario.tipo_usuario.descripcion == 'USUARIO OPERACIONES':

            ordenes_mes = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    pais_id=request.user.pais_id,
                    tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial(),
                    categoria__division__tipo_catalogo__descripcion="SERVICIOS"
                ).\
                exclude(fecha_fin_trabajo__lte='2000-01-01').\
                extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
                extra({'anio': 'Extract(year from fecha_fin_trabajo)'}).\
                order_by('anio', 'mes').\
                values('mes').\
                annotate(suma=Sum('peso_final'))
        else:
            ordenes_mes = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    pais_id=request.user.pais_id,
                    taller=request.user.taller(),
                    categoria__division__tipo_catalogo__descripcion="SERVICIOS"
                ).\
                exclude(fecha_fin_trabajo__lte='2000-01-01').\
                extra({'mes': 'Extract(month from fecha_fin_trabajo)'}).\
                extra({'anio': 'Extract(year from fecha_fin_trabajo)'}).\
                order_by('anio', 'mes').\
                values('mes').\
                annotate(suma=Sum('peso_final'))
        

        # for orden in ordenes_mes:

        fabricaciones_por_mes = []

        for orden in ordenes_mes:
            if orden.get('suma') == None:
                orden['suma'] = float(0)
            orden_tmp = [
                int(orden.get('mes')),
                str(orden.get('suma'))
            ]
            # print((orden_tmp), '****')
            if orden.get('mes') != None:
                fabricaciones_por_mes.append(orden_tmp)

        cont = len(fabricaciones_por_mes)
        if opcion == '1':
            while cont <= 12:
                cont += 1
                orden_tmp = [
                    (cont),
                    str(0.00)
                ]
                fabricaciones_por_mes.append(orden_tmp)
                if cont >= 12:
                    break
        # print(fabricaciones_por_mes, 'lista de la base')
        
        enumerar_meses = [] 
        if opcion == '2':
            enumerar_meses = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
            lista_mes = []
            fabricaciones_por_mes = sorted(fabricaciones_por_mes)
            for mes in fabricaciones_por_mes:
                lista_mes.append(mes[0])
            for mes_lista in enumerar_meses:
                if mes_lista not in lista_mes:
                    orden_tmp = [
                        (mes_lista),
                        str(0.00)
                    ]
                    fabricaciones_por_mes.append(orden_tmp)
            fabricaciones_por_mes = sorted(
                fabricaciones_por_mes,
                key=lambda orden: int(orden[0]),
                reverse=False
            )   
                

        if opcion == '4':
            longitud = len(fabricaciones_por_mes)
            if longitud < 3:
                ultimo_mes = fecha_fin.month
                print(ultimo_mes)
                primer_mes = ultimo_mes - 2
                print(primer_mes)
                print('*****')
                contador_mes = primer_mes
                while contador_mes < ultimo_mes:
                    orden_tmp = [
                        int(contador_mes),
                        str(0.00)
                    ]
                    fabricaciones_por_mes.append(orden_tmp)
                    contador_mes += 1
                    if contador_mes == ultimo_mes:
                        break
                fabricaciones_por_mes = sorted(
                        fabricaciones_por_mes,
                        key=lambda orden: int(orden[0]),
                        reverse=False
                    )


        fabricaciones_por_mes = json.dumps(fabricaciones_por_mes)


        return JsonResponse(fabricaciones_por_mes, safe=False)
    return JsonResponse({}, status = 400)




def cargar_fabricaciones(request):
    if request.is_ajax and request.method == 'GET':
        opcion = request.GET.get("opcion", None)
        f_inicio = request.GET.get("f_inicio", None)
        f_fin = request.GET.get("f_fin", None)
        lista = []

        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        dia_uno_mes_anterior = dia_uno_mes_actual - timedelta(days=1)
        dia_uno_mes_anterior = dia_uno_mes_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        if opcion == '1':
            fecha_inicio = dia_uno_mes_actual 
            fecha_fin = ahora
        if opcion == '2':
            fecha_inicio = dia_uno_mes_anterior
            fecha_fin = dia_uno_mes_actual
        if opcion == '3':
            fecha_inicio = datetime.strptime(f_inicio, "%Y-%m-%d")
            fecha_fin = datetime.strptime(f_fin, "%Y-%m-%d")

        if request.user.tipo_usuario.descripcion == 'USUARIO OPERACIONES':
            ordenes_producto = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                    tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial(),
                    pais=request.user.pais,
                    estado_id=17
                ).\
                values('categoria__permite_solicitud').\
                order_by('categoria__permite_solicitud').\
                annotate(contador=Count('categoria__division__tipo_catalogo__descripcion'))


            ordenes_servicio = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    categoria__division__tipo_catalogo__descripcion='SERVICIOS',
                    tienda__sociedad__grupo_empresarial=request.user.grupo_empresarial(),
                    pais=request.user.pais,
                    estado_id=17
                ).\
                values('categoria__permite_solicitud').\
                order_by('categoria__permite_solicitud').\
                annotate(contador=Count('categoria__division__tipo_catalogo__descripcion'))
        # usuarios de taller
        else:
            ordenes_producto = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    categoria__division__tipo_catalogo__descripcion='PRODUCTOS',
                    taller=request.user.taller(),
                    pais=request.user.pais,
                    estado_id=17
                ).\
                values('categoria__permite_solicitud').\
                order_by('categoria__permite_solicitud').\
                annotate(contador=Count('categoria__division__tipo_catalogo__descripcion'))


            ordenes_servicio = OrdenTrabajo.objects.\
                filter(
                    fecha_fin_trabajo__gte=fecha_inicio,
                    fecha_fin_trabajo__lte=fecha_fin,
                    categoria__division__tipo_catalogo__descripcion='SERVICIOS',
                    taller=request.user.taller(),
                    pais=request.user.pais,
                    estado_id=17
                ).\
                values('categoria__permite_solicitud').\
                order_by('categoria__permite_solicitud').\
                annotate(contador=Count('categoria__division__tipo_catalogo__descripcion'))
        

        lista_catalogo = []
        for p in ordenes_producto:
            
            if p.get('categoria__permite_solicitud'):
                orden_tmp = [
                    'PRODUCTO ESPECIALES',
                    str(p.get('contador')) 
                ]
            else:
                orden_tmp = [
                    'PRODUCTO CATALOGO',
                    str(p.get('contador'))
                ]
            lista_catalogo.append(orden_tmp)

        lista_especiales = []
        for p in ordenes_servicio:
            
            if p.get('categoria__permite_solicitud'):
                orden_tmp = [
                    'SERVICIO ESPECIALES',
                    str(p.get('contador')) 
                ]
            else:
                orden_tmp = [
                    'SERVICIO CATALOGO',
                    str(p.get('contador'))
                ]
            lista_especiales.append(orden_tmp)

        
        lista = lista_catalogo + lista_especiales
        lista = set(map(tuple, lista))
        # ordenar lista
        lista = sorted(
            lista,
            key=lambda orden: int(orden[1]),
            reverse=True
        )

        lista = json.dumps(lista)

        return JsonResponse(lista, safe=False)
    return JsonResponse({}, status = 400)


class CiudadOrdenes(generic.TemplateView):
    template_name = "dsh/estadistica_ciudad_modal.html"
    login_url = "bases:login"

    def get_context_data(self, **kwargs):
        context = super(CiudadOrdenes, self).get_context_data(**kwargs)
        return context

def ciudades_ordenes(request):
    if request.is_ajax and request.method == 'GET':
        opcion = request.GET.get("opcion", None)
        f_inicio = request.GET.get("f_inicio", None)
        f_fin = request.GET.get("f_fin", None)
        anio_actual = datetime.now()
        anio_actual = anio_actual.year
        anio_anterior = (anio_actual) - 1
        dia_uno_anio_actual = datetime(anio_actual, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio = datetime(anio_actual, 12, 31, 11, 59, 59, 00000)
        dia_uno_anio_anterior = datetime(anio_anterior, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio_anterior = datetime(anio_anterior, 12, 31, 11, 59, 59, 00000)


        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        dia_uno_meses_anterior = dia_uno_mes_actual - relativedelta.relativedelta(months=6)
        dia_uno_meses_anterior = dia_uno_meses_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        # print(dia_uno_meses_anterior)

        dia_uno_meses_tres = dia_uno_mes_actual - relativedelta.relativedelta(months=3)
        dia_uno_meses_tres = dia_uno_meses_tres.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )


        if opcion == '1':
            fecha_inicio = dia_uno_anio_actual
            fecha_fin = ultimo_dia_anio
        if opcion == '2':
            fecha_inicio = dia_uno_anio_anterior
            fecha_fin = ultimo_dia_anio_anterior
        if opcion == '3':
            fecha_inicio = dia_uno_meses_anterior
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)
            fecha_fin = fecha_fin.replace(
                hour=11,
                minute=59,
                second=0
                )
        if opcion == '4':
            fecha_inicio = dia_uno_meses_tres
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)


        ciudades_id = Ciudades.objects.filter(pais_id=request.user.pais_id).values('id_ciudad')
        ordenes_ciudad = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__ciudad_id__in=ciudades_id,
                pais=request.user.pais
            ).values('tienda__ciudad__nombre').\
            annotate(contador=Count('id_orden')).\
            order_by('contador')
        lista_ciudades=[]
        for ciudad in ordenes_ciudad:
            orden_ciudad={
                'ciudad': ciudad.get('tienda__ciudad__nombre'),
                'contador': ciudad.get('contador')
            }
            lista_ciudades.append(orden_ciudad)

        lista_ciudades = json.dumps(lista_ciudades)

    return JsonResponse(lista_ciudades, safe=False)


class ZonasOrdenes(generic.TemplateView):
    template_name = "dsh/estadisticas_zona_modal.html"
    login_url = "bases:login"

    def get_context_data(self, **kwargs):
        context = super(ZonasOrdenes, self).get_context_data(**kwargs)
        return context

def zonas_ordenes(request):
    if request.is_ajax and request.method == 'GET':
        opcion = request.GET.get("opcion", None)
        f_inicio = request.GET.get("f_inicio", None)
        f_fin = request.GET.get("f_fin", None)
        anio_actual = datetime.now()
        anio_actual = anio_actual.year
        anio_anterior = (anio_actual) - 1
        dia_uno_anio_actual = datetime(anio_actual, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio = datetime(anio_actual, 12, 31, 11, 59, 59, 00000)
        dia_uno_anio_anterior = datetime(anio_anterior, 1, 1, 00, 00, 00, 10000)
        ultimo_dia_anio_anterior = datetime(anio_anterior, 12, 31, 11, 59, 59, 00000)


        ahora = datetime.now(timezone.utc)
        dia_uno_mes_actual = ahora.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )

        dia_uno_meses_anterior = dia_uno_mes_actual - relativedelta.relativedelta(months=6)
        dia_uno_meses_anterior = dia_uno_meses_anterior.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )
        # print(dia_uno_meses_anterior)

        dia_uno_meses_tres = dia_uno_mes_actual - relativedelta.relativedelta(months=3)
        dia_uno_meses_tres = dia_uno_meses_tres.replace(
            day=1,
            hour=0,
            minute=0,
            second=0
            )


        if opcion == '1':
            fecha_inicio = dia_uno_anio_actual
            fecha_fin = ultimo_dia_anio
        if opcion == '2':
            fecha_inicio = dia_uno_anio_anterior
            fecha_fin = ultimo_dia_anio_anterior
        if opcion == '3':
            fecha_inicio = dia_uno_meses_anterior
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)
            fecha_fin = fecha_fin.replace(
                hour=11,
                minute=59,
                second=0
                )
        if opcion == '4':
            fecha_inicio = dia_uno_meses_tres
            fecha_fin = dia_uno_mes_actual - timedelta(days=1)


        zonas_id = Zonas.objects.filter(grupo_empresarial=request.user.grupo_empresarial()).values('id_zona')

        ordenes_zona = OrdenTrabajo.objects.\
            filter(
                fecha_fin_trabajo__gte=fecha_inicio,
                fecha_fin_trabajo__lte=fecha_fin,
                tienda__zona_id__in=zonas_id,
                pais=request.user.pais
            ).values('tienda__zona__nombre').\
            annotate(contador=Count('id_orden')).\
            order_by('contador')
        lista_zonas=[]
        for zona in ordenes_zona:
            orden_zona={
                'zona': zona.get('tienda__zona__nombre'),
                'contador': zona.get('contador')
            }
            lista_zonas.append(orden_zona)

        lista_zonas = json.dumps(lista_zonas)
        print(lista_zonas)
    return JsonResponse(lista_zonas, safe=False)